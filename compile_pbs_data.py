import os
import re
import sys
import openpyxl
import pandas as pd
from datetime import datetime
from pathlib import Path
from pypdf import PdfReader

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

BASE_DIR = Path(__file__).resolve().parent
PBS_DATA_DIR = BASE_DIR / "pbs_monthly_data"
PROCESSED_DIR = BASE_DIR / "data" / "processed"
OUTPUT_DIR = BASE_DIR / "data" / "output"

CITIES = [
    "Islamabad", "Rawalpindi", "Gujranwala", "Sialkot", "Lahore",
    "Faisalabad", "Sargodha", "Multan", "Bahawalpur", "Karachi",
    "Hyderabad", "Sukkur", "Larkana", "Peshawar", "Bannu",
    "Quetta", "Khuzdar"
]

# 7 Target commodities mirroring exact PBS nomenclature
TARGET_COMMODITY_MAPPING = {
    17: "Pulse Masoor (Washed)",
    18: "Pulse Moong (Washed)",
    19: "Pulse Mash (Washed)",
    20: "Pulse Gram",
    21: "Potatoes",
    22: "Onions",
    23: "Tomatoes"
}

STANDARD_ITEMS_51 = {
    1: ("Wheat Flour Bag", "20 Kg", "Cereals & Grains"),
    2: ("Rice Basmati Broken (Average Quality)", "1 Kg", "Cereals & Grains"),
    3: ("Rice IRRI-6/9 (Sindh/Punjab)", "1 Kg", "Cereals & Grains"),
    4: ("Bread plain (Small Size)", "Each", "Bakery & Confectionery"),
    5: ("Beef with Bone (Average Quality)", "1 Kg", "Meat & Poultry"),
    6: ("Mutton (Average Quality)", "1 Kg", "Meat & Poultry"),
    7: ("Chicken Farm Broiler (Live)", "1 Kg", "Meat & Poultry"),
    8: ("Milk fresh (Un-boiled)", "1 Ltr", "Dairy & Eggs"),
    9: ("Curd (Dahi) Loose", "1 Kg", "Dairy & Eggs"),
    10: ("Powdered Milk NIDO 390 gm Polybag", "Each", "Dairy & Eggs"),
    11: ("Eggs Hen (Farm)", "1 Dozen", "Dairy & Eggs"),
    12: ("Mustard Oil (Average Quality)", "1 Kg", "Oils & Fats"),
    13: ("Cooking Oil DALDA or Other Similar Brand (SN), 5 Litre Tin", "Each", "Oils & Fats"),
    14: ("Vegetable Ghee DALDA/HABIB 2.5 kg Tin", "Each", "Oils & Fats"),
    15: ("Vegetable Ghee DALDA/HABIB or Other superior Quality 1 kg Pouch", "Each", "Oils & Fats"),
    16: ("Bananas (Kela) Local", "1 Dozen", "Fruits"),
    17: ("Pulse Masoor (Washed)", "1 Kg", "Pulses & Legumes"),
    18: ("Pulse Moong (Washed)", "1 Kg", "Pulses & Legumes"),
    19: ("Pulse Mash (Washed)", "1 Kg", "Pulses & Legumes"),
    20: ("Pulse Gram", "1 Kg", "Pulses & Legumes"),
    21: ("Potatoes", "1 Kg", "Vegetables"),
    22: ("Onions", "1 Kg", "Vegetables"),
    23: ("Tomatoes", "1 Kg", "Vegetables"),
    24: ("Sugar Refined", "1 Kg", "Sugar & Sweeteners"),
    25: ("Gur (Average Quality)", "1 Kg", "Sugar & Sweeteners"),
    26: ("Salt Powdered (NATIONAL/SHAN) 800 gm Packet", "Each", "Condiments & Spices"),
    27: ("Chilies Powder NATIONAL 200 gm Packet", "Each", "Condiments & Spices"),
    28: ("Garlic (Lehsun)", "1 Kg", "Vegetables"),
    29: ("Tea Lipton Yellow Label 190 gm Packet", "Each", "Beverages"),
    30: ("Cooked Beef at Average Hotel", "Per Plate", "Food Away from Home"),
    31: ("Cooked Daal at Average Hotel", "Per Plate", "Food Away from Home"),
    32: ("Tea Prepared Ordinary", "Per Cup", "Food Away from Home"),
    33: ("Cigarettes Capstan 20'S Packet", "Each", "Tobacco"),
    34: ("Long Cloth 57\" Gul Ahmed/Al Karam", "1 mtr", "Textiles & Apparel"),
    35: ("Shirting (Average Quality)", "1 mtr", "Textiles & Apparel"),
    36: ("Lawn Printed Gul Ahmed/Al Karam", "1 mtr", "Textiles & Apparel"),
    37: ("Georgette (Average Quality)", "1 mtr", "Textiles & Apparel"),
    38: ("Gents Sandal Bata", "Pair", "Footwear"),
    39: ("Gents Sponge Chappal Bata", "Pair", "Footwear"),
    40: ("Ladies Sandal Bata", "Pair", "Footwear"),
    41: ("Electricity Charges upto 50 Units", "Per Unit", "Energy & Utilities"),
    42: ("Gas Charges upto 3.3719 MMBTU", "MMBTU", "Energy & Utilities"),
    43: ("Firewood Whole", "40 Kg", "Energy & Utilities"),
    44: ("Energy Saver Philips 14 Watt", "Each", "Household Goods"),
    45: ("Sufi Washing Soap 250 gm Cake", "Each", "Personal & Household Care"),
    46: ("Match Box", "Each", "Household Goods"),
    47: ("Petrol Super", "Per Litre", "Fuel & Transport"),
    48: ("Hi-Speed Diesel", "Per Litre", "Fuel & Transport"),
    49: ("LPG 11.67 kg Cylinder", "Each", "Energy & Utilities"),
    50: ("Telephone Call Charges", "Per Minute", "Communication"),
    51: ("Toilet Soap LIFEBUOY 115 gm", "Each", "Personal & Household Care"),
}

def parse_xlsx_file(filepath: Path, ym: str):
    """Parse a single PBS Annexure XLSX file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    
    for r in range(3, min(65, ws.max_row + 1)):
        sno_val = ws.cell(r, 1).value
        if sno_val is None:
            continue
        try:
            sno = int(float(str(sno_val).strip()))
        except (ValueError, TypeError):
            continue
            
        if sno < 1 or sno > 51:
            continue
            
        desc = str(ws.cell(r, 2).value or "").strip()
        unit = str(ws.cell(r, 3).value or "").strip()
        
        city_prices = {}
        for c_idx, city_name in enumerate(CITIES):
            col_num = 4 + c_idx
            c_val = ws.cell(r, col_num).value
            if c_val is not None and str(c_val).strip() != "":
                try:
                    city_prices[city_name] = float(str(c_val).strip().replace(',', ''))
                except ValueError:
                    city_prices[city_name] = None
            else:
                city_prices[city_name] = None
                
        nat_avg = None
        c21_val = ws.cell(r, 21).value
        if c21_val is not None and str(c21_val).strip() != "":
            try:
                nat_avg = float(str(c21_val).strip().replace(',', ''))
            except ValueError:
                pass
                
        if nat_avg is None:
            valid_city_prices = [p for p in city_prices.values() if p is not None]
            if valid_city_prices:
                nat_avg = round(sum(valid_city_prices) / len(valid_city_prices), 2)
                
        item_std_name, item_std_unit, category = STANDARD_ITEMS_51.get(sno, (desc, unit, "General"))
        
        rec = {
            "Year_Month": ym,
            "Date": f"{ym}-01",
            "S_No": sno,
            "Commodity_Raw": desc,
            "Commodity": item_std_name,
            "Target_Commodity": TARGET_COMMODITY_MAPPING.get(sno, None),
            "Category": category,
            "Unit": item_std_unit if item_std_unit else unit,
            "National_Average_Price": nat_avg,
            "Source_File": filepath.name
        }
        rec.update(city_prices)
        records.append(rec)
        
    return records

def parse_pdf_file(filepath: Path, ym: str):
    """Parse a single PBS Annexure PDF file."""
    reader = PdfReader(filepath)
    full_text = ""
    for p in reader.pages:
        full_text += p.extract_text() + "\n"
        
    records = []
    found_snos = set()
    
    for line in full_text.split("\n"):
        line = line.strip()
        m = re.match(r'^(\d{1,2})\s+([A-Za-z].*?)\s+(\d+\.\d{2}.*)$', line)
        if m:
            sno = int(m.group(1))
            if 1 <= sno <= 51 and sno not in found_snos:
                found_snos.add(sno)
                desc = m.group(2).strip()
                nums_str = m.group(3).strip()
                floats = [float(t) for t in re.findall(r'[-+]?\d*\.?\d+', nums_str) if t not in ['-', '+', '.']]
                
                city_prices = {}
                for idx, city_name in enumerate(CITIES):
                    if idx < len(floats):
                        city_prices[city_name] = floats[idx]
                    else:
                        city_prices[city_name] = None
                        
                nat_avg = None
                if len(floats) >= 18:
                    nat_avg = floats[17]
                else:
                    valid_city_prices = [p for p in city_prices.values() if p is not None]
                    if valid_city_prices:
                        nat_avg = round(sum(valid_city_prices) / len(valid_city_prices), 2)
                        
                item_std_name, item_std_unit, category = STANDARD_ITEMS_51.get(sno, (desc, "1 Kg", "General"))
                
                rec = {
                    "Year_Month": ym,
                    "Date": f"{ym}-01",
                    "S_No": sno,
                    "Commodity_Raw": desc,
                    "Commodity": item_std_name,
                    "Target_Commodity": TARGET_COMMODITY_MAPPING.get(sno, None),
                    "Category": category,
                    "Unit": item_std_unit,
                    "National_Average_Price": nat_avg,
                    "Source_File": filepath.name
                }
                rec.update(city_prices)
                records.append(rec)
                
    return records

def main():
    print("=" * 70)
    print("📊 Compiling PBS Monthly Price Datasets (Strict PBS Terminology)")
    print("=" * 70)
    
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    files = sorted([f for f in os.listdir(PBS_DATA_DIR) if f.endswith(('.xlsx', '.pdf'))])
    print(f"[+] Found {len(files)} raw files to compile in {PBS_DATA_DIR.name}/")
    
    all_records = []
    
    for fname in files:
        fpath = PBS_DATA_DIR / fname
        ym_match = re.match(r'^(\d{4}-\d{2})', fname)
        if not ym_match:
            continue
        ym = ym_match.group(1)
        
        if fname.endswith(".xlsx"):
            try:
                recs = parse_xlsx_file(fpath, ym)
                all_records.extend(recs)
                print(f"  [✓] {ym} (Excel): Parsed {len(recs)} commodities")
            except Exception as e:
                print(f"  [-] {ym} (Excel Error): {e}")
        elif fname.endswith(".pdf"):
            try:
                recs = parse_pdf_file(fpath, ym)
                all_records.extend(recs)
                print(f"  [✓] {ym} (PDF):   Parsed {len(recs)} commodities")
            except Exception as e:
                print(f"  [-] {ym} (PDF Error): {e}")
                
    master_df = pd.DataFrame(all_records)
    master_df["Date"] = pd.to_datetime(master_df["Date"])
    master_df = master_df.sort_values(by=["Date", "S_No"]).reset_index(drop=True)
    
    print("\n" + "=" * 70)
    print(f"[+] Master dataset compiled: {len(master_df):,} total monthly observations")
    print(f"[+] Date span: {master_df['Date'].min().strftime('%Y-%m')} to {master_df['Date'].max().strftime('%Y-%m')} ({master_df['Date'].nunique()} distinct months)")
    
    # 1. Save Master 51 Commodities Dataset
    master_csv = PROCESSED_DIR / "master_all_commodities_monthly_prices.csv"
    master_df.to_csv(master_csv, index=False)
    print(f"[✓] Saved: {master_csv} ({len(master_df):,} rows)")
    
    # 2. Filter and Save Target 7 Commodities Dataset
    target_df = master_df[master_df["Target_Commodity"].notnull()].copy()
    target_csv = PROCESSED_DIR / "target_commodities_monthly_prices.csv"
    target_df.to_csv(target_csv, index=False)
    print(f"[✓] Saved: {target_csv} ({len(target_df):,} rows)")
    
    # 3. Create Power BI Ready Historical Prices Table
    pbi_df = target_df[[
        "Date", "Year_Month", "Target_Commodity", "Unit", "National_Average_Price"
    ] + CITIES].copy()
    pbi_df.rename(columns={"Target_Commodity": "Commodity", "National_Average_Price": "Retail_Price_PKR"}, inplace=True)
    
    pbi_long_records = []
    for _, row in pbi_df.iterrows():
        pbi_long_records.append({
            "Date": row["Date"],
            "Year_Month": row["Year_Month"],
            "Commodity": row["Commodity"],
            "Unit": row["Unit"],
            "City": "National Average",
            "Price_PKR": row["Retail_Price_PKR"]
        })
        for city in CITIES:
            if pd.notnull(row[city]):
                pbi_long_records.append({
                    "Date": row["Date"],
                    "Year_Month": row["Year_Month"],
                    "Commodity": row["Commodity"],
                    "Unit": row["Unit"],
                    "City": city,
                    "Price_PKR": row[city]
                })
                
    pbi_long_df = pd.DataFrame(pbi_long_records)
    pbi_csv = OUTPUT_DIR / "historical_commodity_prices.csv"
    pbi_long_df.to_csv(pbi_csv, index=False)
    print(f"[✓] Saved: {pbi_csv} ({len(pbi_long_df):,} rows - Power BI ready)")
    
    print("\n" + "=" * 70)
    print("📈 TARGET COMMODITIES PRICE SUMMARY (Latest vs Earliest available)")
    print("=" * 70)
    for comm in sorted(target_df["Target_Commodity"].unique()):
        sub = target_df[target_df["Target_Commodity"] == comm].sort_values("Date")
        earliest_p = sub.iloc[0]["National_Average_Price"]
        latest_p = sub.iloc[-1]["National_Average_Price"]
        earliest_d = sub.iloc[0]["Year_Month"]
        latest_d = sub.iloc[-1]["Year_Month"]
        pct_change = ((latest_p - earliest_p) / earliest_p) * 100
        print(f"  * {comm:<25s}: {earliest_d} = Rs {earliest_p:>6.2f}  ──►  {latest_d} = Rs {latest_p:>6.2f}  ({pct_change:>+6.1f}%)")
    print("=" * 70)

if __name__ == "__main__":
    main()
