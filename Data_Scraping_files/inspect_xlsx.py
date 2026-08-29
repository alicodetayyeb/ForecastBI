"""
Script to inspect structure and sample rows of downloaded PBS Excel (.xlsx) Annexure files.
"""

import os
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
PBS_DIR = BASE_DIR / "pbs_monthly_data"
xlsx_files = sorted([f for f in os.listdir(PBS_DIR) if f.endswith(".xlsx")])

print(f"Found {len(xlsx_files)} XLSX files in {PBS_DIR.name}:")

for fname in xlsx_files[:3]:
    fpath = PBS_DIR / fname
    print("\n" + "=" * 60)
    print(f"File: {fname}")
    wb = openpyxl.load_workbook(fpath, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sheetname in wb.sheetnames[:2]:
        ws = wb[sheetname]
        print(f"--- Sheet: {sheetname} (max_row: {ws.max_row}, max_col: {ws.max_column}) ---")
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else "" for c in range(1, min(15, ws.max_column + 1))]
            if any(row_vals):
                print(f"Row {r:2d}: {row_vals}")
